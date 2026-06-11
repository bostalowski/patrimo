"""Append PEE transactions and ensure the corresponding FCPE assets exist.

Idempotent: re-running won't duplicate assets or transactions.
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import openpyxl


_RAW = os.environ.get("EXCEL_PATH")
if not _RAW:
    raise SystemExit(
        "EXCEL_PATH is not set. Run with: EXCEL_PATH=... python3 scripts/add_pee_transactions.py",
    )
EXCEL = Path(_RAW).expanduser().resolve()


ASSETS_TO_ENSURE: list[dict[str, str]] = [
    {
        "id": "FCPE-IMPACT-ISR",
        "label": "Impact ISR Performance (FCPE)",
        "type": "FCPE",
        "isin": "QS0004088926",
        "ticker": "",
        "source": "investir",
        "param": "QS0004088926",
        "currency": "EUR",
    },
]


TRANSACTIONS_TO_APPEND: list[dict] = [
    {
        "date": datetime(2026, 5, 27),
        "type": "ACHAT",
        "compte": "PEE",
        "compte_destination": "",
        "actif": "FCPE-IMPACT-ISR",
        "quantite": 19.9821,
        "prix_unitaire": 61.75383,
        "devise": "EUR",
        "frais": 0,
        "frais_devise": "EUR",
        "notes": "",
    },
    {
        "date": datetime(2026, 5, 27),
        "type": "ACHAT",
        "compte": "PEE",
        "compte_destination": "",
        "actif": "FCPE-MIROVA-INTL",
        "quantite": 7.5138,
        "prix_unitaire": 38.04582,
        "devise": "EUR",
        "frais": 0,
        "frais_devise": "EUR",
        "notes": "",
    },
    {
        "date": datetime(2026, 5, 27),
        "type": "ACHAT",
        "compte": "PEE",
        "compte_destination": "",
        "actif": "FCPE-WATER",
        "quantite": 8.4395,
        "prix_unitaire": 29.62246,
        "devise": "EUR",
        "frais": 0,
        "frais_devise": "EUR",
        "notes": "",
    },
    {
        "date": datetime(2026, 5, 28),
        "type": "ACHAT",
        "compte": "PEE",
        "compte_destination": "",
        "actif": "FCPE-IMPACT-ISR",
        "quantite": 0.2603,
        "prix_unitaire": 61.73375,
        "devise": "EUR",
        "frais": 0,
        "frais_devise": "EUR",
        "notes": "",
    },
]


PEE_ALLOCATION_CATEGORY = "PEE diversifié"


def ensure_assets(ws) -> int:
    headers = {cell.value: cell.column for cell in ws[1]}
    existing_ids = {
        row[headers["ID"] - 1].value
        for row in ws.iter_rows(min_row=2)
        if row[headers["ID"] - 1].value
    }
    added = 0
    for asset in ASSETS_TO_ENSURE:
        if asset["id"] in existing_ids:
            continue
        ws.append(
            [
                asset["id"],
                asset["label"],
                asset["type"],
                asset["isin"],
                asset["ticker"],
                asset["source"],
                asset["param"],
                asset["currency"],
            ]
        )
        added += 1
    return added


def transaction_key(row_values: tuple) -> tuple:
    date, tx_type, compte, _dest, actif, quantite, prix, *_ = row_values
    return (date, tx_type, compte, actif, quantite, prix)


def append_transactions(ws) -> int:
    existing = {
        transaction_key(row)
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }
    added = 0
    for tx in TRANSACTIONS_TO_APPEND:
        key = (
            tx["date"],
            tx["type"],
            tx["compte"],
            tx["actif"],
            tx["quantite"],
            tx["prix_unitaire"],
        )
        if key in existing:
            continue
        ws.append(
            [
                tx["date"],
                tx["type"],
                tx["compte"],
                tx["compte_destination"],
                tx["actif"],
                tx["quantite"],
                tx["prix_unitaire"],
                tx["devise"],
                tx["frais"],
                tx["frais_devise"],
                tx["notes"],
            ]
        )
        last_row = ws.max_row
        ws.cell(row=last_row, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=last_row, column=6).number_format = "0.########"
        ws.cell(row=last_row, column=7).number_format = "#,##0.########"
        ws.cell(row=last_row, column=9).number_format = "#,##0.##"
        added += 1
    return added


def update_allocation(ws) -> bool:
    headers = {cell.value: cell.column for cell in ws[1]}
    cat_col = headers["Catégorie"]
    actifs_col = headers["Actifs (séparés par virgule)"]
    new_ids = {a["id"] for a in ASSETS_TO_ENSURE if a["type"] == "FCPE"}
    if not new_ids:
        return False
    for row in ws.iter_rows(min_row=2):
        if row[cat_col - 1].value != PEE_ALLOCATION_CATEGORY:
            continue
        cell = row[actifs_col - 1]
        current = [s.strip() for s in str(cell.value or "").split(",") if s.strip()]
        merged = current + [i for i in new_ids if i not in current]
        if merged == current:
            return False
        cell.value = ",".join(merged)
        return True
    return False


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL)
    assets_added = ensure_assets(wb["Actifs"])
    txs_added = append_transactions(wb["Transactions"])
    alloc_updated = update_allocation(wb["Allocation cible"])
    wb.save(EXCEL)
    print(
        f"Assets added: {assets_added}\n"
        f"Transactions added: {txs_added}\n"
        f"Allocation updated: {alloc_updated}"
    )


if __name__ == "__main__":
    main()
