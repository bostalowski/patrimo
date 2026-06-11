"""Replace the two BTC RETRAIT entries (which only recorded network fees) with
proper TRANSFERT entries Kraken → Ledger.

Numbers come from the user's Ledger Live data:
  Transfer 1 (2025-11-27): received 0.00983748 BTC, network fee 0.00002211 BTC
  Transfer 2 (2026-06-01): received 0.00950080 BTC, network fee 0.00001110 BTC
"""

from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import openpyxl

_RAW = os.environ.get("EXCEL_PATH")
if not _RAW:
    raise SystemExit(
        "EXCEL_PATH is not set. Run with: EXCEL_PATH=... python3 scripts/fix_btc_transfers.py",
    )
EXCEL = Path(_RAW).expanduser().resolve()

TRANSFERS: list[dict] = [
    {
        "date": datetime(2025, 11, 27),
        "kraken_withdrawn": 0.000265,
        "kraken_fee": 0.000015,
        "received_at_ledger": 0.00025,
        "network_fee": None,
        "notes": "Transfert Kraken→Ledger (27/11/2025 18:23, chiffres Kraken)",
    },
    {
        "date": datetime(2025, 11, 27),
        "kraken_withdrawn": 0.00985248,
        "kraken_fee": 0.000015,
        "received_at_ledger": 0.00983748,
        "network_fee": 0.00002211,
        "notes": "Transfert Kraken→Ledger (27/11/2025 20:34, chiffres Kraken)",
    },
    {
        "date": datetime(2026, 6, 1),
        "kraken_withdrawn": 0.0095158,
        "kraken_fee": 0.000015,
        "received_at_ledger": 0.00950080,
        "network_fee": 0.00001110,
        "notes": "Transfert Kraken→Ledger (chiffres Kraken)",
    },
]


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Transactions"]
    headers = {cell.value: cell.column for cell in ws[1]}

    rows_to_delete: list[int] = []
    for row in ws.iter_rows(min_row=2):
        type_cell = row[headers["Type"] - 1]
        actif_cell = row[headers["Actif"] - 1]
        if type_cell.value in ("RETRAIT", "TRANSFERT") and actif_cell.value == "BTC":
            rows_to_delete.append(row[0].row)

    print(f"Removing {len(rows_to_delete)} legacy BTC row(s) at: {rows_to_delete}")
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)

    for t in TRANSFERS:
        if t["kraken_withdrawn"] is not None:
            quantity_out = t["kraken_withdrawn"]
            fee = t["kraken_fee"]
        else:
            quantity_out = t["received_at_ledger"] + t["network_fee"]
            fee = t["network_fee"]
        ws.append(
            [
                t["date"],
                "TRANSFERT",
                "Kraken",
                "Ledger",
                "BTC",
                round(quantity_out, 8),
                None,
                "EUR",
                round(fee, 8),
                "BTC",
                t["notes"],
            ]
        )
        last_row = ws.max_row
        ws.cell(row=last_row, column=headers["Date"]).number_format = "yyyy-mm-dd"
        ws.cell(row=last_row, column=headers["Quantité"]).number_format = "0.########"
        ws.cell(row=last_row, column=headers["Frais"]).number_format = "0.########"

    wb.save(EXCEL)
    print(f"Added {len(TRANSFERS)} TRANSFERT row(s).")


if __name__ == "__main__":
    main()
