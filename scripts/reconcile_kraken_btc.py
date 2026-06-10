"""Reconcile Kraken BTC ledger against screenshots (June 2026).

Operations:
- Remove 2 phantom DIVIDENDE rows on 2025-12-03 (actually 03/01/2026 conversions)
- Remove the duplicate DIVIDENDE on 2025-12-28 (qty 0.00000285)
- Move ACHAT 30 EUR from 2025-12-29 to 2025-11-29
- Add 3 missing rewards: 09/11, 11/11, 28/11
- Add 2 missing conversions on 03/01/2026 (reuse old precise qty)
"""

from datetime import datetime
from pathlib import Path

import openpyxl

EXCEL_PATH = Path("data/Investissement.xlsx")
SHEET = "Transactions"


def find_row(ws, headers, predicate):
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        record = {name: values[idx] for name, idx in headers.items()}
        if predicate(record):
            return row[0].row, record
    return None, None


def approx(a, b, tol=1e-9):
    if a is None or b is None:
        return False
    return abs(a - b) < tol


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET]
    headers = {c.value: i for i, c in enumerate(ws[1])}

    to_delete = []
    rescued_qty = []

    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        rec = {name: values[idx] for name, idx in headers.items()}
        if rec["Actif"] != "BTC" or rec["Compte"] != "Kraken":
            continue
        date = rec["Date"]
        type_ = rec["Type"]
        qty = rec["Quantité"]

        if (
            type_ == "DIVIDENDE"
            and date == datetime(2025, 12, 3)
            and qty in (0.00003804, 0.00003719)
        ):
            to_delete.append(row[0].row)
            rescued_qty.append(qty)
            print(f"  - DELETE row {row[0].row}: 2025-12-03 DIVIDENDE {qty:.8f} (will be moved to 2026-01-03)")

        elif (
            type_ == "DIVIDENDE"
            and date == datetime(2025, 12, 28)
            and approx(qty, 0.00000285)
        ):
            to_delete.append(row[0].row)
            print(f"  - DELETE row {row[0].row}: 2025-12-28 DIVIDENDE {qty:.8f} (duplicate)")

    for row_idx in sorted(to_delete, reverse=True):
        ws.delete_rows(row_idx, 1)

    row_idx, rec = find_row(
        ws,
        headers,
        lambda r: (
            r["Actif"] == "BTC"
            and r["Compte"] == "Kraken"
            and r["Type"] == "ACHAT"
            and r["Date"] == datetime(2025, 12, 29)
            and approx(r["Quantité"], 0.00037323)
        ),
    )
    if row_idx:
        ws.cell(row=row_idx, column=headers["Date"] + 1).value = datetime(2025, 11, 29)
        ws.cell(row=row_idx, column=headers["Date"] + 1).number_format = "yyyy-mm-dd"
        print(f"  ~ MOVE   row {row_idx}: ACHAT 30€ 2025-12-29 → 2025-11-29")

    additions = [
        (datetime(2025, 11, 9), "DIVIDENDE", "Kraken", None, "BTC", 0.00000485, 0, "EUR", 0, "EUR",
         "Kraken Earn reward 0,43 € (qty estimée d'après cours BTC)"),
        (datetime(2025, 11, 11), "DIVIDENDE", "Kraken", None, "BTC", 0.00001920, 0, "EUR", 0, "EUR",
         "Kraken Earn reward 1,73 €"),
        (datetime(2025, 11, 28), "DIVIDENDE", "Kraken", None, "BTC", 0.00000280, 0, "EUR", 0, "EUR",
         "Kraken Earn reward 0,22 €"),
        (datetime(2026, 1, 3), "DIVIDENDE", "Kraken", None, "BTC", rescued_qty[0] if rescued_qty else 0.00003804, 0, "EUR", 0, "EUR",
         "Conversion 0,00111 ETH → BTC (récompense d'origine ETH staking)"),
        (datetime(2026, 1, 3), "DIVIDENDE", "Kraken", None, "BTC", rescued_qty[1] if len(rescued_qty) > 1 else 0.00003719, 0, "EUR", 0, "EUR",
         "Conversion autre actif → BTC (récompense d'origine)"),
    ]

    for row_data in additions:
        ws.append(list(row_data))
        last = ws.max_row
        ws.cell(row=last, column=headers["Date"] + 1).number_format = "yyyy-mm-dd"
        ws.cell(row=last, column=headers["Quantité"] + 1).number_format = "0.########"
        ws.cell(row=last, column=headers["Prix unitaire"] + 1).number_format = "#,##0.##"
        date_str = row_data[0].strftime("%Y-%m-%d")
        print(f"  + ADD    {date_str} DIVIDENDE Kraken BTC {row_data[5]:.8f}")

    wb.save(EXCEL_PATH)
    print(f"\nDone: deleted {len(to_delete)}, added {len(additions)}")


if __name__ == "__main__":
    main()
