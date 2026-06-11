"""Migrate the legacy single-sheet Investissement.xlsx into the new 3-sheet schema.

Source : the original Google-Sheets-exported xlsx (single "Mouvements" tab).
Target : a clean xlsx with Transactions / Actifs / Comptes tabs.

Run :
    python3 scripts/migrate_excel.py \
        --src /Users/bastien.ostalowski/Downloads/Investissement.xlsx \
        --dst "$HOME/Library/CloudStorage/GoogleDrive-<email>/Mon Drive/Investissement.xlsx"
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


ASSET_MAPPING: dict[str, dict[str, str]] = {
    "Bitcoin": {
        "id": "BTC",
        "label": "Bitcoin",
        "type": "CRYPTO",
        "isin": "",
        "ticker": "BTC",
        "source": "coingecko",
        "param": "bitcoin",
        "currency": "EUR",
    },
    "S&P 500": {
        "id": "SP500",
        "label": "S&P 500 (à préciser)",
        "type": "ETF",
        "isin": "",
        "ticker": "",
        "source": "manual",
        "param": "",
        "currency": "EUR",
    },
    "Orange": {
        "id": "ORANGE",
        "label": "Orange",
        "type": "ACTION",
        "isin": "FR0000133308",
        "ticker": "ORA.PA",
        "source": "yahoo",
        "param": "ORA.PA",
        "currency": "EUR",
    },
    "Amundi PEA profil equilibré": {
        "id": "FCPE-AMUNDI-EQUILIBRE",
        "label": "Amundi PEA Profil Équilibré",
        "type": "FCPE",
        "isin": "",
        "ticker": "",
        "source": "manual",
        "param": "",
        "currency": "EUR",
    },
    "Ishare MSCI World": {
        "id": "ISHARE-MSCI-WORLD",
        "label": "iShares MSCI World (PEA)",
        "type": "ETF",
        "isin": "",
        "ticker": "",
        "source": "manual",
        "param": "",
        "currency": "EUR",
    },
    "Amundi MSCI World": {
        "id": "AMUNDI-MSCI-WORLD",
        "label": "Amundi MSCI World (PEA)",
        "type": "ETF",
        "isin": "",
        "ticker": "",
        "source": "manual",
        "param": "",
        "currency": "EUR",
    },
    "Amundi emerging ESG": {
        "id": "AMUNDI-EMERGING-ESG",
        "label": "Amundi MSCI Emerging Markets ESG (PEA)",
        "type": "ETF",
        "isin": "",
        "ticker": "",
        "source": "manual",
        "param": "",
        "currency": "EUR",
    },
    "Mirova actions internationnales": {
        "id": "FCPE-MIROVA-INTL",
        "label": "Mirova Actions Internationales",
        "type": "FCPE",
        "isin": "",
        "ticker": "",
        "source": "manual",
        "param": "",
        "currency": "EUR",
    },
    "Thematic water": {
        "id": "FCPE-WATER",
        "label": "Thematic Water",
        "type": "FCPE",
        "isin": "",
        "ticker": "",
        "source": "manual",
        "param": "",
        "currency": "EUR",
    },
    "Intérêts": {
        "id": "INTERETS",
        "label": "Intérêts cash",
        "type": "CASH",
        "isin": "",
        "ticker": "",
        "source": "manual",
        "param": "",
        "currency": "EUR",
    },
}

ACCOUNT_MAPPING: dict[str, dict[str, str]] = {
    "CTO": {"id": "CTO", "label": "Compte-titres ordinaire", "type": "BROKER", "envelope": "CTO"},
    "PEA": {"id": "PEA", "label": "PEA courtier", "type": "BROKER", "envelope": "PEA"},
    "PEE": {"id": "PEE", "label": "Natixis PEE", "type": "EPARGNE_SALARIALE", "envelope": "PEE"},
    "Kraken": {"id": "Kraken", "label": "Kraken", "type": "EXCHANGE_CRYPTO", "envelope": "CTO"},
    "Ledger": {"id": "Ledger", "label": "Ledger (cold wallet)", "type": "WALLET_CRYPTO", "envelope": "CTO"},
}


@dataclass
class Transaction:
    date: datetime
    type: str
    compte: str
    compte_destination: str
    actif: str
    quantite: float
    prix_unitaire: float | None
    devise: str
    frais: float
    frais_devise: str
    notes: str


def map_action_to_type(action: str, libelle: str, prix: float, quantite: float) -> str:
    if action == "Vente":
        return "VENTE"
    if action == "Dividende":
        return "DIVIDENDE"
    if action == "Intérêts":
        return "INTERET"
    if action == "Cashin":
        return "DEPOT"
    if action == "Achat":
        if libelle == "Bitcoin" and prix == 0:
            return "DIVIDENDE" if quantite >= 0 else "RETRAIT"
        return "ACHAT"
    raise ValueError(f"Unknown action: {action!r}")


def migrate(src: Path, dst: Path) -> None:
    src_wb = openpyxl.load_workbook(src, data_only=False)
    ws = src_wb["Mouvements"]

    transactions: list[Transaction] = []
    seen_assets: set[str] = set()
    seen_accounts: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        date, libelle, action, direction, prix, quantite, _g, frais, _i, *_ = row
        if date is None or libelle is None or action is None:
            continue

        libelle_key = str(libelle).strip()
        if libelle_key not in ASSET_MAPPING:
            raise ValueError(f"Unmapped asset libellé: {libelle_key!r}")
        asset = ASSET_MAPPING[libelle_key]

        direction_key = str(direction).strip()
        if direction_key not in ACCOUNT_MAPPING:
            raise ValueError(f"Unmapped account direction: {direction_key!r}")
        account = ACCOUNT_MAPPING[direction_key]

        prix_value = float(prix) if prix is not None else 0.0
        qty_value = float(quantite) if quantite is not None else 0.0
        frais_value = abs(float(frais)) if frais is not None else 0.0

        tx_type = map_action_to_type(str(action).strip(), libelle_key, prix_value, qty_value)

        qty_out = abs(qty_value)
        prix_out: float | None = prix_value if tx_type not in ("TRANSFERT",) else None
        if tx_type in ("DIVIDENDE", "RETRAIT") and prix_value == 0:
            prix_out = 0.0

        transactions.append(
            Transaction(
                date=date,
                type=tx_type,
                compte=account["id"],
                compte_destination="",
                actif=asset["id"],
                quantite=qty_out,
                prix_unitaire=prix_out,
                devise="EUR",
                frais=frais_value,
                frais_devise="EUR",
                notes="",
            )
        )

        seen_assets.add(libelle_key)
        seen_accounts.add(direction_key)

    transactions.sort(key=lambda t: (t.date, t.actif))

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    write_transactions(out_wb, transactions)
    write_actifs(out_wb, seen_assets)
    write_comptes(out_wb, seen_accounts | {"Ledger"})

    dst.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(dst)
    print(f"Wrote {len(transactions)} transactions to {dst}")


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


def style_header(ws, header_count: int) -> None:
    for col in range(1, header_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def autosize(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_transactions(wb, transactions: list[Transaction]) -> None:
    ws = wb.create_sheet("Transactions")
    headers = [
        "Date",
        "Type",
        "Compte",
        "Compte destination",
        "Actif",
        "Quantité",
        "Prix unitaire",
        "Devise",
        "Frais",
        "Frais devise",
        "Notes",
    ]
    ws.append(headers)
    style_header(ws, len(headers))
    autosize(ws, [12, 12, 14, 18, 26, 14, 14, 8, 10, 12, 28])
    ws.freeze_panes = "A2"

    for tx in transactions:
        ws.append(
            [
                tx.date,
                tx.type,
                tx.compte,
                tx.compte_destination,
                tx.actif,
                tx.quantite,
                tx.prix_unitaire,
                tx.devise,
                tx.frais,
                tx.frais_devise,
                tx.notes,
            ]
        )

    for row_idx in range(2, len(transactions) + 2):
        ws.cell(row=row_idx, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=row_idx, column=6).number_format = "0.########"
        ws.cell(row=row_idx, column=7).number_format = "#,##0.##"
        ws.cell(row=row_idx, column=9).number_format = "#,##0.##"


def write_actifs(wb, seen_libelles: set[str]) -> None:
    ws = wb.create_sheet("Actifs")
    headers = ["ID", "Libellé", "Type", "ISIN", "Ticker", "Source prix", "Param source", "Devise"]
    ws.append(headers)
    style_header(ws, len(headers))
    autosize(ws, [24, 36, 10, 16, 14, 14, 22, 8])
    ws.freeze_panes = "A2"

    for libelle, a in ASSET_MAPPING.items():
        if libelle not in seen_libelles:
            continue
        ws.append([a["id"], a["label"], a["type"], a["isin"], a["ticker"], a["source"], a["param"], a["currency"]])


def write_comptes(wb, seen_directions: set[str]) -> None:
    ws = wb.create_sheet("Comptes")
    headers = ["ID", "Libellé", "Type", "Enveloppe"]
    ws.append(headers)
    style_header(ws, len(headers))
    autosize(ws, [16, 28, 22, 14])
    ws.freeze_panes = "A2"

    for direction, c in ACCOUNT_MAPPING.items():
        if direction not in seen_directions:
            continue
        ws.append([c["id"], c["label"], c["type"], c["envelope"]])


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--src", required=True, type=Path)
    parser.add_argument("--dst", required=True, type=Path)
    args = parser.parse_args()
    migrate(args.src, args.dst)


if __name__ == "__main__":
    main()
