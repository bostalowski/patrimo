"""Switch FCPE assets to the new 'investir' source (Investir.lesechos.fr scraper)."""

from pathlib import Path
import openpyxl

EXCEL = Path("data/Investissement.xlsx")

UPDATES: dict[str, dict[str, str]] = {
    "FCPE-MIROVA-INTL": {"source": "investir", "param": "QS0004036743"},
    "FCPE-WATER": {"source": "investir", "param": "QS0004037725"},
}


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb["Actifs"]
    headers = {cell.value: cell.column for cell in ws[1]}
    updated = 0
    for row in ws.iter_rows(min_row=2):
        id_ = row[headers["ID"] - 1].value
        if id_ in UPDATES:
            row[headers["Source prix"] - 1].value = UPDATES[id_]["source"]
            row[headers["Param source"] - 1].value = UPDATES[id_]["param"]
            updated += 1
    wb.save(EXCEL)
    print(f"Updated {updated} assets to investir source")


if __name__ == "__main__":
    main()
