"""Rename asset IDs and fill in ISIN/Ticker/Source columns after user confirmation.

Idempotent : will only rename if the old IDs are still present, no-op otherwise.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl


EXCEL = Path("data/Investissement.xlsx")

ID_RENAMES: dict[str, str] = {
    "AMUNDI-MSCI-WORLD": "DCAM",
    "ISHARE-MSCI-WORLD": "WPEA",
    "AMUNDI-EMERGING-ESG": "PLEM",
    "SP500": "AMUNDI-SP500",
}

ASSET_UPDATES: dict[str, dict[str, object]] = {
    "DCAM": {
        "label": "Amundi PEA Monde (MSCI World) UCITS ETF Acc",
        "type": "ETF",
        "isin": "FR001400U5Q4",
        "ticker": "DCAM.PA",
        "source": "yahoo",
        "param": "DCAM.PA",
    },
    "WPEA": {
        "label": "iShares MSCI World Swap PEA UCITS ETF EUR Acc",
        "type": "ETF",
        "isin": "IE0002XZSHO1",
        "ticker": "WPEA.PA",
        "source": "yahoo",
        "param": "WPEA.PA",
    },
    "PLEM": {
        "label": "Amundi PEA Emergent EMEA ESG Transition UCITS ETF Acc",
        "type": "ETF",
        "isin": "FR0011440478",
        "ticker": "PLEM.PA",
        "source": "yahoo",
        "param": "PLEM.PA",
    },
    "AMUNDI-SP500": {
        "label": "Amundi S&P 500 Swap UCITS ETF EUR Acc",
        "type": "ETF",
        "isin": "LU1681048804",
        "ticker": "500.PA",
        "source": "yahoo",
        "param": "500.PA",
    },
    "FCPE-WATER": {
        "label": "Thematic Water (FCPE)",
        "type": "FCPE",
        "isin": "QS0004037725",
        "ticker": "",
        "source": "manual",
        "param": "",
    },
    "FCPE-MIROVA-INTL": {
        "label": "Mirova Actions Internationales (FCPE)",
        "type": "FCPE",
        "isin": "QS0004036743",
        "ticker": "",
        "source": "manual",
        "param": "",
    },
}


def rename_transactions(ws) -> int:
    count = 0
    actif_col = None
    for cell in ws[1]:
        if cell.value == "Actif":
            actif_col = cell.column
            break
    if actif_col is None:
        raise RuntimeError("Cannot find 'Actif' column in Transactions sheet")
    for row in ws.iter_rows(min_row=2):
        cell = row[actif_col - 1]
        if cell.value in ID_RENAMES:
            cell.value = ID_RENAMES[cell.value]
            count += 1
    return count


def update_actifs(ws) -> tuple[int, int]:
    renamed = 0
    updated = 0
    headers = {cell.value: cell.column for cell in ws[1]}
    for row in ws.iter_rows(min_row=2):
        id_cell = row[headers["ID"] - 1]
        if id_cell.value in ID_RENAMES:
            id_cell.value = ID_RENAMES[id_cell.value]
            renamed += 1
        update = ASSET_UPDATES.get(id_cell.value)
        if update:
            for key, col_name in [
                ("label", "Libellé"),
                ("type", "Type"),
                ("isin", "ISIN"),
                ("ticker", "Ticker"),
                ("source", "Source prix"),
                ("param", "Param source"),
            ]:
                row[headers[col_name] - 1].value = update[key]
            updated += 1
    return renamed, updated


def update_allocation(ws) -> int:
    count = 0
    headers = {cell.value: cell.column for cell in ws[1]}
    col = headers["Actifs (séparés par virgule)"]
    for row in ws.iter_rows(min_row=2):
        cell = row[col - 1]
        if cell.value is None:
            continue
        ids = [s.strip() for s in str(cell.value).split(",")]
        new_ids = [ID_RENAMES.get(i, i) for i in ids]
        if new_ids != ids:
            cell.value = ",".join(new_ids)
            count += 1
    return count


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL)
    tx_count = rename_transactions(wb["Transactions"])
    renamed, updated = update_actifs(wb["Actifs"])
    alloc_count = update_allocation(wb["Allocation cible"])
    wb.save(EXCEL)
    print(
        f"Transactions renamed: {tx_count}\n"
        f"Actifs renamed: {renamed}, updated: {updated}\n"
        f"Allocation rows updated: {alloc_count}"
    )


if __name__ == "__main__":
    main()
